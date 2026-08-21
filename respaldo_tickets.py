#!/usr/bin/env python3
"""
Respaldo de TODOS los tickets a un archivo Excel (.xlsx).

Se conecta directo a la misma base de datos MongoDB que usa el backend
(lee MONGODB_URI y COLLECTION_NAME de backend/.env), en modo solo lectura:
no modifica ni borra nada. Genera un archivo con toda la información de
todos los tickets (canjeados o no), coloreando cada fila según el día
(hora de Ecuador) en que se hizo el canje — las filas sin canjear quedan
sin color, porque no tienen un día de canje que colorear.

Uso:
    python respaldo_tickets.py
    python respaldo_tickets.py --salida respaldo_20-08.xlsx
    python respaldo_tickets.py --env otra/ruta/.env

Instalar dependencias primero:
    pip install -r requirements.txt
"""

import argparse
import os
import sys
import unicodedata
from datetime import datetime, timedelta
from pathlib import Path

from dotenv import load_dotenv
from pymongo import MongoClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

RAIZ = Path(__file__).resolve().parent

# America/Guayaquil no tiene horario de verano, así que un offset fijo
# alcanza (no hace falta zoneinfo/tzdata para esto)
ECUADOR_OFFSET = timedelta(hours=-5)

COLUMNAS = [
    'Name', 'Last Name', 'Email', 'Ticket', 'Seat', 'Transaction ID', 'Ticket ID',
    'Cedula', 'Last four', 'RETIRA', 'NÚMERO DE CEDULA', 'CELULAR', 'Responsable', 'Fecha'
]

LAST4_SIN_DATO = 'Cash / Comprado en punto de venta'

# Paleta de colores pastel, uno por día (si hay más días que colores, se repite)
PALETA_DIAS = [
    'D9EAD3', 'CFE2F3', 'FFF2CC', 'F4CCCC', 'D9D2E9', 'FCE5CD',
    'D0E0E3', 'EAD1DC', 'B6D7A8', 'A4C2F4', 'FFE599', 'EA9999'
]


def normalizar(texto):
    """Quita tildes y pasa a minúsculas, para comparar nombres de columna
    sin importar la variante exacta (con/sin tilde, con/sin espacio final)."""
    texto = str(texto or '')
    sin_tildes = ''.join(
        c for c in unicodedata.normalize('NFD', texto) if not unicodedata.combining(c)
    )
    return sin_tildes.lower().strip()


def obtener_cedula(ticket):
    """Cédula del comprador: el nombre de la columna varía según cómo se
    importó el CSV ('Numero de Cedula:', 'Número de Cédula:', etc.), así que
    se busca cualquier clave que contenga 'cedula' ya normalizada."""
    clave = next((k for k in ticket.keys() if 'cedula' in normalizar(k)), None)
    return str(ticket.get(clave, '') or '').strip() if clave else ''


def obtener_last4(ticket):
    valor = str(ticket.get('Last4/PayPal Email', '') or '').strip()
    return valor or LAST4_SIN_DATO


def formatear_retira(ticket):
    quien = ticket.get('quienRetira') or ''
    if quien == 'Otro':
        parentesco = ticket.get('parentesco') or 'N/A'
        quien_otro = ticket.get('quienOtro') or ''
        return f'Otro ({parentesco}: {quien_otro})'
    return quien or 'N/A'


def a_hora_ecuador(fecha_utc):
    # pymongo entrega datetimes "naive" en UTC (así los guarda MongoDB)
    if fecha_utc is None:
        return None
    return fecha_utc + ECUADOR_OFFSET


def formatear_fecha(fecha_ecuador):
    if fecha_ecuador is None:
        return ''
    return fecha_ecuador.strftime('%d/%m/%Y %H:%M')


def conectar(env_path):
    load_dotenv(env_path)
    uri = os.environ.get('MONGODB_URI')
    coleccion = os.environ.get('COLLECTION_NAME')

    if not uri:
        sys.exit(f'No se encontró MONGODB_URI (revisá {env_path} o exportala en el entorno)')
    if not coleccion:
        sys.exit(f'No se encontró COLLECTION_NAME (revisá {env_path} o exportala en el entorno)')

    cliente = MongoClient(uri)
    db = cliente.get_default_database()
    if db is None:
        sys.exit('El MONGODB_URI no incluye el nombre de la base de datos (falta "/nombreDB" en la URL)')

    return db, coleccion


def obtener_nombres_usuarios(db):
    """Mapa {ObjectId: nombre} de la colección Usuarios, para resolver el
    responsable del canje sin hacer una consulta por cada ticket."""
    return {u['_id']: u.get('nombre', '') for u in db['Usuarios'].find({}, {'nombre': 1})}


def construir_filas(tickets, nombres_usuarios):
    """Devuelve una lista de (fila_dict, dia_ecuador) — el día se usa
    después para elegir el color de cada fila en el Excel."""
    filas = []
    for ticket in tickets:
        fecha_ecuador = a_hora_ecuador(ticket.get('fechaCanje'))
        responsable_id = ticket.get('usuarioCanje') or ticket.get('usuarioResponsable')
        responsable = nombres_usuarios.get(responsable_id, '') if responsable_id else ''

        fila = {
            'Name': ticket.get('First Name', ''),
            'Last Name': ticket.get('Last Name', ''),
            'Email': ticket.get('Email', ''),
            'Ticket': ticket.get('Ticket', ''),
            'Seat': ticket.get('Seat', ''),
            'Transaction ID': ticket.get('Transaction ID', ''),
            'Ticket ID': ticket.get('Ticket ID', ''),
            'Cedula': obtener_cedula(ticket),
            'Last four': obtener_last4(ticket),
            'RETIRA': formatear_retira(ticket),
            'NÚMERO DE CEDULA': ticket.get('cedulaQuienRetira', ''),
            'CELULAR': ticket.get('celular', ''),
            'Responsable': responsable,
            'Fecha': formatear_fecha(fecha_ecuador),
        }
        filas.append((fila, fecha_ecuador.date() if fecha_ecuador else None))
    return filas


def armar_excel(filas, ruta_salida):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Canjeados'

    ws.append(COLUMNAS)
    for celda in ws[1]:
        celda.font = Font(bold=True, color='FFFFFF')
        celda.fill = PatternFill('solid', fgColor='434343')
        celda.alignment = Alignment(horizontal='center')
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(COLUMNAS))}1'

    # Un color por día, en orden cronológico
    dias_ordenados = sorted({dia for _, dia in filas if dia is not None})
    color_por_dia = {dia: PALETA_DIAS[i % len(PALETA_DIAS)] for i, dia in enumerate(dias_ordenados)}

    for fila, dia in filas:
        ws.append([fila[col] for col in COLUMNAS])
        if dia is not None:
            relleno = PatternFill('solid', fgColor=color_por_dia[dia])
            for celda in ws[ws.max_row]:
                celda.fill = relleno

    # Ancho de columna aproximado según el contenido más largo
    for i, col in enumerate(COLUMNAS, start=1):
        largo_max = max([len(col)] + [len(str(f[col])) for f, _ in filas])
        ws.column_dimensions[get_column_letter(i)].width = min(largo_max + 2, 45)

    # Hoja aparte con la leyenda: qué color corresponde a cada día
    if dias_ordenados:
        leyenda = wb.create_sheet('Leyenda de colores')
        leyenda.append(['Día (hora Ecuador)', 'Cantidad canjeada'])
        leyenda['A1'].font = Font(bold=True)
        leyenda['B1'].font = Font(bold=True)

        conteo_por_dia = {}
        for _, dia in filas:
            if dia is not None:
                conteo_por_dia[dia] = conteo_por_dia.get(dia, 0) + 1

        for dia in dias_ordenados:
            fila_num = leyenda.max_row + 1
            leyenda.append([dia.strftime('%d/%m/%Y'), conteo_por_dia[dia]])
            for celda in leyenda[fila_num]:
                celda.fill = PatternFill('solid', fgColor=color_por_dia[dia])

        leyenda.column_dimensions['A'].width = 22
        leyenda.column_dimensions['B'].width = 20

    wb.save(ruta_salida)


def main():
    parser = argparse.ArgumentParser(
        description='Respaldo de todos los tickets a Excel, coloreado por día de canje'
    )
    parser.add_argument(
        '--env', default=str(RAIZ / 'backend' / '.env'),
        help='Ruta al .env con MONGODB_URI y COLLECTION_NAME (default: backend/.env)'
    )
    parser.add_argument(
        '--salida', default=None,
        help='Nombre del archivo de salida (default: respaldo_tickets_YYYY-MM-DD_HHMM.xlsx)'
    )
    args = parser.parse_args()

    print('Conectando a MongoDB...')
    db, nombre_coleccion = conectar(args.env)
    print(f'Conectado. Colección activa: {nombre_coleccion}')

    print('Descargando todos los tickets...')
    # Sin filtro por 'canjeado': se trae toda la base. Ordenados por fecha de
    # canje para que las filas coloreadas queden agrupadas por día; los que
    # todavía no se canjearon (sin fechaCanje) quedan sin colorear.
    tickets = list(db[nombre_coleccion].find({}).sort('fechaCanje', 1))
    print(f'  {len(tickets)} ticket(s) encontrados en total')

    if not tickets:
        print('No hay tickets en la base, no se genera archivo')
        return

    nombres_usuarios = obtener_nombres_usuarios(db)
    filas = construir_filas(tickets, nombres_usuarios)

    salida = args.salida or f'respaldo_tickets_{datetime.now().strftime("%Y-%m-%d_%H%M")}.xlsx'
    print(f'Generando {salida}...')
    armar_excel(filas, salida)

    dias = sorted({dia for _, dia in filas if dia is not None})
    print(f'Listo: {len(filas)} fila(s), {len(dias)} día(s) distinto(s) coloreados')
    print(f'Archivo: {Path(salida).resolve()}')


if __name__ == '__main__':
    main()
